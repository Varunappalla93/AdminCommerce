import openpyxl

def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

def getcolcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_column

def readData(file,sheetname,rowno,colno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno,column=colno).value

def writeData(file,sheetname,rowno,colno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno,column=colno).value=data
    workbook.save(file)